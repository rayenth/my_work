from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
import os



def ishere(path):
    if (os.path.exists("GPU-DATA.xlsx")):
        wb=load_workbook("GPU-DATA.xlsx")
        return wb
    else :
        wb=Workbook()
        return wb




def sanitize_sheet_name(name):
    return re.sub(r'[\\/:*?"[<>|]', '_', name)
def make_excel(dic,text):
    dt=datetime.now()
    fdt=dt.strftime("%Y-%m-%d %H:%M:%S")
    wb=ishere("GPU-DATA.xlsx")
    sheet_name = f"result for {text} in {fdt}"
    sanitized_sheet_name = sanitize_sheet_name(sheet_name)
    ws = wb.create_sheet( sanitized_sheet_name)

    ws.append(list(dic.keys()))
    for i in range (0,len(list(dic["Memory"][1]))):
        l=[]
        for j in list(dic.keys()): 
            l.append(dic[j][1][i])
            
        ws.append(l)

    apply_text_wrapping(ws)
    adjust_column_widths(ws, padding=5)

#making a font for the header
    for column_num in range (1,len(dic.keys())+1):
        col_letter=get_column_letter(column_num)
        ws[f"{col_letter}1"].font=Font(bold=True,color="ffa000")


    wb.save("GPU-DATA.xlsx")


#this will show the entire text in the cell
def apply_text_wrapping(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


#and this will desplay the text in a single ligne instead of desplaying it in a wrapped manner 
def adjust_column_widths(ws, padding=2):
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = max_length + padding
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = adjusted_width
    